#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
daily_timesheet.py
严格按员工清单顺序输出，并增加“部门”列
"""
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

RAW_FILE    = 'yunxiao_20250903_095231_794.xlsx'
STAFF_FILE  = '员工清单.xlsx'
OUTPUT_FILE = f'工时统计-{datetime.now():%Y%m%d%H%M%S}.xlsx'

RED_FONT = Font(color='FF0000')

def load_staff():
    """返回 [(姓名, 部门), ...] 保持清单原始顺序"""
    ws = load_workbook(STAFF_FILE).active
    header = [c.value for c in ws[1]]
    name_idx = header.index('人员名称')
    dept_idx = header.index('部门')
    return [(row[name_idx], row[dept_idx]) for row in ws.iter_rows(min_row=2, values_only=True)]

def load_raw():
    """返回 dict：姓名 -> {'工时': x, '项数': y}"""
    ws = load_workbook(RAW_FILE)['工时投入排名']
    return {row[0]: {'工时': row[1] or 0, '项数': row[2] or 0}
            for row in ws.iter_rows(min_row=3, values_only=True)}

def main():
    staff = load_staff()          # 顺序与部门
    raw   = load_raw()            # 已填数据

    wb = load_workbook(RAW_FILE)
    ws = wb['工时投入排名']

    # 清空旧数据（保留表头）
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.font  = Font()

    # 写入新表头（第1、2、3列已存在；第4列插入“部门”）
    ws['D1'] = '部门'

    # 按员工清单顺序写入
    for idx, (name, dept) in enumerate(staff, start=3):
        rec = raw.get(name, {'工时': 0, '项数': 0})
        hours, items = rec['工时'], rec['项数']

        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=hours)
        ws.cell(row=idx, column=3, value=items)
        ws.cell(row=idx, column=4, value=dept)

        # 标红：工时不足 7 或未填工时
        if hours < 7:
            ws.cell(row=idx, column=1).font = RED_FONT   # 姓名
            ws.cell(row=idx, column=2).font = RED_FONT   # 工时

    wb.save(OUTPUT_FILE)
    print('已生成：', OUTPUT_FILE)

if __name__ == '__main__':
    main()