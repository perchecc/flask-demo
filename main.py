from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有来源，生产环境请指定具体域名
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 临时文件存储目录
TEMP_DIR = "temp_files"
os.makedirs(TEMP_DIR, exist_ok=True)

RED_FONT = Font(color='FF0000')
DEFAULT_FONT = Font()

def load_staff(staff_file):
    """从上传的员工清单文件加载数据，明确匹配表头：人员名称、部门"""
    try:
        wb = load_workbook(staff_file)
        ws = wb.active
        
        # 查找包含'人员名称'和'部门'的表头行
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=5):  # 检查前5行
            values = [str(c.value).strip() if c.value is not None else "" for c in row]
            if '人员名称' in values and '部门' in values:
                header_row = row
                break
        
        if not header_row:
            raise ValueError("未找到包含'人员名称'和'部门'的表头行")
            
        # 获取列索引
        values = [str(c.value).strip() if c.value is not None else "" for c in header_row]
        name_idx = values.index('人员名称')
        dept_idx = values.index('部门')
        
        # 从表头行的下一行开始读取数据
        start_row = header_row[0].row + 1
        staff_data = []
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
            if len(row) > name_idx and row[name_idx]:  # 确保姓名不为空
                dept_value = row[dept_idx] if len(row) > dept_idx else ""
                staff_data.append((row[name_idx], dept_value))
        
        return staff_data
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"员工清单文件处理错误: {str(e)}")

def load_raw(raw_file):
    """从上传的原始工时文件加载数据，明确匹配表头：人员名称、登记工时(小时)、工作项数"""
    try:
        wb = load_workbook(raw_file, data_only=True)  # 读取单元格的值而非公式
        
        # 尝试查找名为"工时投入排名"的工作表，如果找不到则使用第一个工作表
        if '工时投入排名' in wb.sheetnames:
            ws = wb['工时投入排名']
        else:
            ws = wb.active
        
        # 查找包含明确表头的行
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=5):  # 检查前5行
            values = [str(c.value).strip() if c.value is not None else "" for c in row]
            if '人员名称' in values and '登记工时(小时)' in values and '工作项数' in values:
                header_row = row
                break
        
        if not header_row:
            raise ValueError("未找到包含'人员名称'、'登记工时(小时)'和'工作项数'的表头行")
            
        # 获取列索引
        values = [str(c.value).strip() if c.value is not None else "" for c in header_row]
        name_idx = values.index('人员名称')
        hours_idx = values.index('登记工时(小时)')
        items_idx = values.index('工作项数')
        
        # 从表头行的下一行开始收集数据
        raw_data = {}
        start_row = header_row[0].row + 1
        
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
            if len(row) > name_idx and row[name_idx]:  # 确保姓名不为空
                name = str(row[name_idx]).strip()  # 标准化姓名
                # 处理工时
                hours = 0.0
                if len(row) > hours_idx and row[hours_idx] is not None:
                    try:
                        hours = float(row[hours_idx])
                    except (ValueError, TypeError):
                        hours = 0.0
                # 处理项数
                items = 0
                if len(row) > items_idx and row[items_idx] is not None:
                    try:
                        items = int(row[items_idx])
                    except (ValueError, TypeError):
                        items = 0
                raw_data[name] = {'工时': hours, '项数': items}
        
        return raw_data
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"原始工时文件处理错误: {str(e)}")

@app.post("/process-timesheet")
async def process_timesheet(
    staff_file: UploadFile = File(..., description="员工清单Excel文件（表头：人员名称、部门）"),
    raw_file: UploadFile = File(..., description="原始工时Excel文件（表头：人员名称、登记工时(小时)、工作项数）")
):
    """处理工时统计并返回结果文件"""
    try:
        # 保存上传的临时文件
        staff_temp_path = os.path.join(TEMP_DIR, staff_file.filename)
        raw_temp_path = os.path.join(TEMP_DIR, raw_file.filename)
        
        with open(staff_temp_path, "wb") as f:
            f.write(await staff_file.read())
        with open(raw_temp_path, "wb") as f:
            f.write(await raw_file.read())

        # 加载数据
        staff = load_staff(staff_temp_path)
        raw_data = load_raw(raw_temp_path)

        # 创建输出文件
        output_filename = f'工时统计-{datetime.now():%Y%m%d%H%M%S}.xlsx'
        output_path = os.path.join(TEMP_DIR, output_filename)
        
        # 创建新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "工时投入排名"
        
        # 写入表头（第一行）
        ws['A1'] = '人员名称'
        ws['B1'] = '工时'
        ws['C1'] = '项数'
        ws['D1'] = '部门'
        
        # 设置表头样式（加粗）
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 按员工清单顺序写入数据，从第二行开始（解决第二行空行问题）
        for idx, (name, dept) in enumerate(staff, start=2):  # 从第2行开始写入数据
            # 查找匹配的记录
            matched_rec = raw_data.get(str(name).strip(), {'工时': 0.0, '项数': 0})
            hours, items = matched_rec['工时'], matched_rec['项数']

            # 写入数据
            name_cell = ws.cell(row=idx, column=1, value=name)
            hour_cell = ws.cell(row=idx, column=2, value=hours)
            item_cell = ws.cell(row=idx, column=3, value=items)
            dept_cell = ws.cell(row=idx, column=4, value=dept)

            # 设置单元格格式
            name_cell.font = DEFAULT_FONT
            hour_cell.font = DEFAULT_FONT
            item_cell.font = DEFAULT_FONT
            dept_cell.font = DEFAULT_FONT

            # 工时小于7小时标红
            if hours < 7:
                name_cell.font = RED_FONT
                hour_cell.font = RED_FONT

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15

        # 保存新工作簿
        wb.save(output_path)

        # 返回生成的文件
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")
    finally:
        # 清理临时文件
        for path in [staff_temp_path, raw_temp_path]:
            if os.path.exists(path):
                os.remove(path)

@app.get("/")
async def read_root():
    return {"message": "工时统计API服务，请使用POST /process-timesheet上传文件"}
    